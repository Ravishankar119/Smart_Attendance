from tkinter import ttk
import cv2
import face_recognition
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os
import threading
import tkinter as tk
from tkinter import messagebox, scrolledtext
from PIL import Image, ImageTk
from pathlib import Path

# Constants
KNOWN_FACES_DIR = 'photos'
EXCEL_FILE = 'attendance.xlsx'
CAMERA_SOURCE = 0
CHECKOUT_THRESHOLD = timedelta(minutes=1)

# Load known faces
def get_face_encoding(image_path):
    image = face_recognition.load_image_file(image_path)
    encodings = face_recognition.face_encodings(image)
    return encodings[0] if encodings else None

known_face_encodings = []
known_face_names = []

for filename in os.listdir(KNOWN_FACES_DIR):
    if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
        name = os.path.splitext(filename)[0]
        encoding = get_face_encoding(os.path.join(KNOWN_FACES_DIR, filename))
        if encoding is not None:
            known_face_encodings.append(encoding)
            known_face_names.append(name)
        else:
            print(f"❌ No face found in {filename}. Skipped.")

# Setup Excel file
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Date', 'Check-in Time', 'Check-out Time'])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

attendance_records = {}
face_timers = {}

def is_real_person(face_encoding, frame, face_location):
    # Placeholder for anti-spoofing or liveness detection
    return True

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition Attendance System")
        self.video_capture = None
        self.running = False
        
        # ========== BACKGROUND IMAGE ==========
        try:
            # Load and resize image
            #self.bg_image = Image.open("")  # Replace with your image path
            # Replace line 65 with:
            image_path = Path("C:\\Users\DELL\\Desktop\\Smart_Attendance\\photos\\rise.jpg")  # Forward slashes work
            self.bg_image = Image.open(image_path)

            self.bg_image = self.bg_image.resize(
                (root.winfo_screenwidth(), root.winfo_screenheight()),
                Image.LANCZOS
            )
            self.bg_photo = ImageTk.PhotoImage(self.bg_image)
            
            # Create background label
            self.background_label = tk.Label(root, image=self.bg_photo)
            self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        except Exception as e:
            print(f"⚠️ Background error: {e}")
            # Fallback to solid color if image fails
            self.background_label = tk.Label(root, bg='gray20')
            self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        # ========== END BACKGROUND ==========

        # ========== ADD THIS SECTION ==========
        # Create orange horizontal bar frame
        self.heading_frame = tk.Frame(root, bg="crimson")
        self.heading_frame.pack(fill="x")  # First widget at top

        # Add label inside the orange frame
        self.heading_label = tk.Label(
            self.heading_frame,
            text="SMART ATTENDANCE",
            font=("Arial", 20, "bold"),
            fg="white",
            bg="crimson"
        )
        self.heading_label.pack(pady=10)

    
        # ========== END ADDED SECTION =========
        

        # GUI Elements
        
        self.video_label = tk.Label(root)
        self.video_label.pack()

        self.log_text = scrolledtext.ScrolledText(root, width=60, height=10, state='disabled')
        self.log_text.pack(pady=10)

        self.btn_frame = tk.Frame(root)
        self.btn_frame.pack()

        self.start_btn = tk.Button(
    self.btn_frame, 
    text="Open camera", 
    command=self.start_attendance,
    bg="light blue",
    fg="black",
    activebackground="light blue",  
    activeforeground="black"       
)
        self.start_btn.grid(row=0, column=0, padx=10)
        self.stop_btn = tk.Button(
    self.btn_frame, 
    text="Stop Attendance", 
    command=self.stop_attendance, 
    state='disabled',
    bg="light green",
    fg="white",
    activebackground="light green",
    activeforeground="white"
)
        self.stop_btn.grid(row=0, column=1, padx=10)

        
        # New Quit button
        self.quit_btn = tk.Button(
            self.btn_frame, 
            text="Quit", 
            command=self.quit_app,
            bg="#ff4444",  # Red color
            fg="white",
            activebackground="#ff0000",
            activeforeground="white"
        )
        self.quit_btn.grid(row=0, column=2, padx=10)

        

    # Add this new method
    def quit_app(self):
        self.stop_attendance()  # Stop camera if running
        self.root.destroy()  # Close the application
#-----------------------------------------------------

    def log(self, message):
        self.log_text.configure(state='normal')
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.configure(state='disabled')
        self.log_text.see(tk.END)

    def start_attendance(self):
        if self.running:
            self.log("Attendance already running.")
            return
        self.video_capture = cv2.VideoCapture(CAMERA_SOURCE)
        if not self.video_capture.isOpened():
            messagebox.showerror("Error", "Cannot open video source.")
            return

        self.running = True
        self.start_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.log("✅ Attendance system started.")
        self.process_frame()

    def stop_attendance(self):
        if not self.running:
            self.log("Attendance is not running.")
            return
        self.running = False
        if self.video_capture:
            self.video_capture.release()
            self.video_capture = None
        self.video_label.config(image='')
        self.start_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        wb.save(EXCEL_FILE)
        self.log("🔴 Attendance system stopped and data saved.")

    def process_frame(self):
        if not self.running:
            return

        ret, frame = self.video_capture.read()
        if not ret:
            self.log("⚠️ Failed to grab frame.")
            self.stop_attendance()
            return

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        now = datetime.now()
        date_str = now.strftime("%d-%m-%y")

        for face_encoding, face_location in zip(face_encodings, face_locations):
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances) if len(face_distances) > 0 else -1

            name = "Unknown"
            if best_match_index >= 0 and face_distances[best_match_index] < 0.45:
                name = known_face_names[best_match_index]

            if name != "Unknown" and is_real_person(face_encoding, frame, face_location):
                top, right, bottom, left = [v * 4 for v in face_location]

                if name in face_timers:
                    elapsed = now - face_timers[name]
                else:
                    face_timers[name] = now
                    elapsed = timedelta(seconds=0)

                # Attendance logic
                if name not in attendance_records:
                    attendance_records[name] = {"date": date_str, "checkin": now, "checkout": None}
                    ws.append([name, date_str, now.strftime("%H:%M:%S"), ""])
                    wb.save(EXCEL_FILE)
                    self.log(f"🟢 Check-in recorded: {name} at {now.strftime('%H:%M:%S')}")

                elif attendance_records[name]["checkout"] is None and \
                     now - attendance_records[name]["checkin"] > CHECKOUT_THRESHOLD:
                    attendance_records[name]["checkout"] = now
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value == name and row[1].value == date_str:
                            row[3].value = now.strftime("%H:%M:%S")
                            break
                    wb.save(EXCEL_FILE)
                    self.log(f"🔁 Check-out recorded: {name} at {now.strftime('%H:%M:%S')}")

                # Draw rectangle and timer
                timer_text = f"{str(elapsed).split('.')[0]}"
                cv2.putText(frame, timer_text, (left, bottom + 30), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 0), 2)

            # Draw rectangle and name
            top, right, bottom, left = [v * 4 for v in face_location]
            color = (0, 0, 255) if name == "Unknown" else (0, 255, 0)
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)

        # Convert frame to ImageTk for Tkinter
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        imgtk = ImageTk.PhotoImage(image=img)
        self.video_label.imgtk = imgtk
        self.video_label.configure(image=imgtk)

        # Call this function again after 10 ms
        self.root.after(10, self.process_frame)
        
        #---------------------------------------------------------------+++++++
        

    if cv2.waitKey(1) & 0xFF == ord('q'):
          print("🔴 Exiting...")

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.protocol("WM_DELETE_WINDOW", app.stop_attendance)  # Ensure cleanup on close
    root.mainloop()
